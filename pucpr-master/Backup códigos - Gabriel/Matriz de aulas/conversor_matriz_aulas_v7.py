# -*- coding: Windows-1252 -*-

from tkinter import Tk
from tkinter.filedialog import askopenfilename
from numpy import nan
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Ajusta as informa��es da tabela de aulas por per�odo
def ajusta_info(df_aulas):
    cont = 1 # Contador do per�odo
    linha = 0 # Contador de linha
    # Insere o n�mero do per�odo na coluna criada
    # Tira os valores desnecess�rios das linhas
    nome_coluna = [] # Cria uma lista contendo o nome de cada coluna
    for c in df_aulas.columns:
        nome_coluna.append(c)
    for dado in df_aulas['Ordem']:
        if (dado == f'{cont+1}� PER�ODO'):
            cont+=1
        if (str(dado)!='nan'):
            df_aulas['Per�odo'][linha] = str(int(cont))
            if (str(dado)=='Ordem' or str(dado)=='TOTAL' or str(dado)==f'{cont}� PER�ODO'):
                df_aulas['Per�odo'][linha] = nan
                coluna = 0
                while coluna < 11:
                    df_aulas[nome_coluna[coluna]][linha] = nan
                    coluna+=1
        linha+=1
    return df_aulas

# Verifica quais colunas possuem dados a serem validados
# Apaga as colunas desnecess�rias/n�o nomeadas
def ajusta_colunas(df):
    colunas = []
    drop_colunas = []
    contador = 0
    for c in df.head():
        # Feito desse modo para assegurar que a coluna ser� encontrada
        if (c != f'Unnamed: {contador-1}' and 
            c != f'Unnamed: {contador}' and 
            c != f'Unnamed: {contador+1}'):
            colunas.append(str(c))
        else:
            drop_colunas.append(str(c))
        contador+=1
    for dc in range(len(drop_colunas)):
        df.drop(drop_colunas[dc],inplace=True,axis=1)
    return df

# Insere os dados conforme o template
def insere_dados(df_para, df_aulas):
    for col_aulas in df_aulas.head():
        for col_para in df_para.head():
            dados_col_aula = []
            cont_dados_aulas = 0
            if (' '+col_aulas == col_para or
                col_aulas == col_para or
                col_aulas+' ' == col_para or
                (col_aulas == 'CRED' and col_para == 'Cr�dito')):
                for dados_aula in df_aulas[col_aulas]:
                    dados_col_aula.append(str(dados_aula))
                    cont_dados_aulas+=1
                for i in range(len(dados_col_aula)):
                    try:
                        # Se o dado inserido puder ser transformado em int, faz a convers�o
                        df_para[col_para][i] = int(dados_col_aula[i].split('.')[0])
                    except:
                        # Caso contr�rio, � armazenado como String
                        df_para[col_para][i] = dados_col_aula[i]
    return df_para

# Transforma os valores n�mericos, necess�rios, em int
def converte_dtype(df_para):
    for col_para in df_para.head():
        dado_para = []
        if (col_para == 'AT' or
            col_para == 'AP' or
            col_para == 'TOTAL' or
            col_para == 'Cr�dito' or
            col_para == 'HA' or
            col_para == 'HR' or
            col_para == 'HR EAD' or
            col_para == 'CH (HR) Extensionista' or
            col_para == 'N�mero de Vagas' or
            col_para == 'Vagas Totais' or
            col_para == 'MT' or
            col_para == 'MP' or
            col_para == 'CH Docente Te�rica' or
            col_para == 'CH Docente Pr�tica' or
            col_para == 'CH Docente Total'):
            for dp in df_para[col_para]:
                dado_para.append(dp)
            for t_dp in range(len(dado_para)):
                try:
                    df_para[col_para][t_dp] = int(dado_para[t_dp])
                except:
                    continue
    return df_para

# Atentar a forma como a fun��o � escrita: Deve ser feita em ingl�s
# TODO: Verificar se todas as fun��es est�o sendo inseridas
# Insere as fun��es nas c�lulas
def ajusta_funcoes(df_para):
    inicio = 14
    fim = 6
    for col_para in df_para.head():
        # Faltam algumas colunas para colocar
        # Essas inser��es de fun��o n�o est�o fucionando >> Provavelmente pq a tabela 2 est� sendo identificada dps 
        # e est� sobreescrevendo os dados, ent�o a inser��o da fun��o funciona, s� n�o aparece
        # TODO: Ajustar os dados da tabela 2
        #if col_para == 'Escola / Campus':
        #    for i in range(len(df_para[col_para])):
        #        if (((len(df_para[col_para])-i)) == 0):
        #            df_para[col_para][i] = f'=SUM(D{i+(inicio-2)}+M{i+(inicio-fim)})'
        #elif col_para == 'Curso':
        #        if ((((len(df_para[col_para])-i) - fim)) == -2):
        #            df_para[col_para][i] = f'=D{i+inicio}*20'
        #        if ((((len(df_para[col_para])-i) - fim)) == -1):
        #            df_para[col_para][i] = f'=F{i+inicio}'
        #        if ((((len(df_para[col_para])-i) - fim)) == 0):
        #            df_para[col_para][i] = f'=(E{i+(inicio-1)}+N{i+(inicio-fim)})'
        #elif col_para == 'Per�odo':
        #        if ((((len(df_para[col_para])-i) - fim)) == -2):
        #            df_para[col_para][i] = f'=D{i+inicio}*15'
        #        if ((((len(df_para[col_para])-i) - fim)) == 0):
        #            df_para[col_para][i] = f'=(E{i+(inicio-1)}+O{i+(inicio-fim)}+F{i-1})'
        if col_para == 'Escola / Campus':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$I$9'
        elif col_para == 'Curso':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$D$9'
        elif col_para == 'TOTAL':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=SUM(J{i+inicio}:K{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,L{inicio}:L{i+(inicio-1)})'
        elif col_para == 'Cr�dito':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=L{i+inicio}'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,M{inicio}:M{i+(inicio-1)})'
        elif col_para == 'HA':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = F'=IF(AND(x{i+inicio}="Est�gio",Y{i+inicio}="Externo"),O{i+inicio},L{i+inicio}*20)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,N{inicio}:N{i+(inicio-1)})'
        elif col_para == 'HR':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=L{i+inicio}*15'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,O{inicio}:O{i+(inicio-1)})'
        elif col_para == 'HR EAD': 
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=IF(X{i+inicio}="EAD",M{i+inicio},0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,P{inicio}:P{i+(inicio-1)})'
        elif col_para == 'CH (HR) Extensionista': # TODO: Falta a refer�ncia correta
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=#REF!'
                    # f'=SUM(H{i+inicio}:I{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = '=#REF!'
                    # f'=SUBTOTAL(109,Q{inicio}:Q{i+(inicio-1)})'
        elif col_para == 'N�mero de Vagas':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=M9'
        elif col_para == 'Vagas Totais':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=IFERROR(VLOOKUP(F{i+inicio},Lista!$AA$2:$AC$13,2,0),"-")'
        elif col_para == 'CH Docente Te�rica':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=ROUNDUP(IF(ISBLANK(AA{i+inicio}),(IFERROR((S{i+inicio}/T{i+inicio})*J{i+inicio},0)),0),0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AC{inicio}:AC{i+(inicio-1)})'
        elif col_para == 'CH Docente Pr�tica':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=ROUNDUP(IF(ISBLANK(AA{i+inicio}),(IFERROR((S{i+inicio}/U{i+inicio})*K{i+inicio},0)),0),0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AD{inicio}:AD{i+(inicio-1)})'
        elif col_para == 'CH Docente Total':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=SUM(AC{i+inicio}:AD{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AE{inicio}:AE{i+(inicio-1)})'
    return df_para

def validacao_dados(load_df_para):

    return load_df_para

def insere_template(load_df_para, load_template):
    fd_para_pagina_origem = load_df_para["Sheet1"]
    fd_para_pagina_destino = load_template['CRIA��O DE MATRIZ PRESENCIAL']

    for i in range(1, fd_para_pagina_origem.max_row+1):
        for j in range(1, fd_para_pagina_origem.max_column+1):
            try:
                fd_para_pagina_destino.cell(row=i+12, column=j+2).value = fd_para_pagina_origem.cell(row=i+12, column=j+2).value
            except:
                continue
    load_template.save('matriz_template.xlsx') # A matriz inserida no template

# Fun��o principal
def main():
    # Sele��o de arquivo dos aulas
    print("Selecione a planilha de aulas")
    Tk().withdraw()
    planilha_aulas = askopenfilename(
        filetypes=[('Arquivo excel', '.xlsx')],
        title='Selecione a planilha de aulas')
    # Mostra qual foi o arquivo selecionado das aulas
    print("Arquivo selecionado: " + str(planilha_aulas).split('/')[-1])
    
    # Selecionando a planilha PARA
    print("Selecione a planilha PARA")
    Tk().withdraw()
    planilha_para = askopenfilename(
        filetypes=[('Arquivo excel', '.xlsx')],
        title='Selecione a planilha PARA')
    # Mostra qual foi o arquivo selecionado PARA
    print("Arquivo selecionado: " + str(planilha_para).split('/')[-1])

    ''' -------------------------------------------------------------------------------
        Leitura das p�ginas do arquivo
        ------------------------------------------------------------------------------- '''
    
    print("Lendo arquivo")
    
    # Planilha de aulas � armazenado na vari�vel
    df_aulas = pd.read_excel(planilha_aulas, dtype=str, header=1)

    # planilha PARA e armazenada em uma variavel, selecionando apartir do header 12
    df_para = pd.read_excel(planilha_para, dtype=str, header=12, index_col=1, sheet_name=0)
    # df_para2 = pd.read_excel(planilha_para, header=115, index_col=1, sheet_name=0) # Era para pegar a tabela 2 para trabalhar

    df_para = ajusta_colunas(df_para) # Faz os ajustes necess�rios nas colunas
    # df_para2 = ajusta_colunas(df_para2)

    df_aulas.insert(loc=0, column='Per�odo', value=nan) # Cria uma coluna para colocar o n�mero do per�odo
    df_aulas = ajusta_info(df_aulas) # Ajusta as informa��es e linhas da tabela de aulas
    
    # Remove os nan das c�lulas vazias
    df_aulas.dropna(inplace=True, how='all') # remove a linha, quando todas as c�lulas estiverem vazias
    df_aulas.drop_duplicates(inplace=True) # remove duplicadas, no caso, os espa�os em branco e os t�tulos
    df_aulas = df_aulas.fillna("")
    
    df_para.drop_duplicates(inplace=True) # remove duplicadas, no caso, os espa�os em branco e os t�tulos
    df_para = df_para.fillna("")
    
    df_para = insere_dados(df_para, df_aulas) # Pega os dados do dataframe de alunos e insere no template
    df_para = ajusta_funcoes(df_para) # TODO: Inserir as fun��es no local correto
    df_para = converte_dtype(df_para)

    '''------------------------------------------------------------------------------------------------'''
    
    df_para.to_excel('template_matriz.xlsx', index=False, startrow=12, startcol=2) # Cria a planilha utilizando o template PARA
    # df_para2.to_excel('template_matriz2.xlsx', index=False) # Isso era para mexer na tabela 2

    load_df_para = load_workbook('template_matriz.xlsx') # Criar o arquivo com o nome do curso
    # load_df_para = validacao_dados(load_df_para)
    load_template = load_workbook(planilha_para)
    
    insere_template(load_df_para, load_template)

if __name__ == '__main__':
    # Chamada da fun��o main
    main()
