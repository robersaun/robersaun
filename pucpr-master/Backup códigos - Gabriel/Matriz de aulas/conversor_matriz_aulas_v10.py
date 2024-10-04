# -*- coding: Windows-1252 -*-

from tkinter import *
from tkinter.filedialog import askopenfilename
from numpy import nan
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Ajusta as informa��es da tabela de aulas por per�odo
def ajusta_info(df_aulas):
    cont = 1 # Contador do per�odo
    linha = 0 # Contador de linha
    # Insere o n�mero do per�odo na coluna criada
    # Tira os valores desnecess�rios das linhas
    nome_coluna = [] # Cria uma lista contendo o nome de cada coluna
    for c in df_aulas.columns:
        nome_coluna.append(c)
    for dado in df_aulas['Ordem']:
        if (dado == f'{cont+1}� PER�ODO'):
            cont+=1
        if (str(dado)!='nan'):
            df_aulas['Per�odo'][linha] = str(int(cont))
            if (str(dado)=='Ordem' or str(dado)=='TOTAL' or str(dado)==f'{cont}� PER�ODO'):
                df_aulas['Per�odo'][linha] = nan
                coluna = 0
                while coluna < 11:
                    df_aulas[nome_coluna[coluna]][linha] = nan
                    coluna+=1
        linha+=1
    return df_aulas

# Verifica quais colunas possuem dados a serem validados
# Apaga as colunas desnecess�rias/n�o nomeadas
def ajusta_colunas(df):
    colunas = []
    drop_colunas = []
    contador = 0
    for c in df.head():
        # Feito desse modo para assegurar que a coluna ser� encontrada
        if (c != f'Unnamed: {contador-1}' and 
            c != f'Unnamed: {contador}' and 
            c != f'Unnamed: {contador+1}'):
            colunas.append(str(c))
        else:
            drop_colunas.append(str(c))
        contador+=1
    for dc in range(len(drop_colunas)):
        df.drop(drop_colunas[dc],inplace=True,axis=1)
    return df

# Insere os dados conforme o template
def insere_dados(df_para, df_aulas):
    for col_aulas in df_aulas.head():
        for col_para in df_para.head():
            dados_col_aula = []
            cont_dados_aulas = 0
            if (' '+col_aulas == col_para or
                col_aulas == col_para or
                col_aulas+' ' == col_para or
                (col_aulas == 'CRED' and col_para == 'Cr�dito')):
                for dados_aula in df_aulas[col_aulas]:
                    dados_col_aula.append(str(dados_aula))
                    cont_dados_aulas+=1
                for i in range(len(dados_col_aula)):
                    try:
                        # Se o dado inserido puder ser transformado em int, faz a convers�o
                        df_para[col_para][i] = int(dados_col_aula[i].split('.')[0])
                    except:
                        # Caso contr�rio, � armazenado como String
                        df_para[col_para][i] = dados_col_aula[i]
    return df_para

# Transforma os valores n�mericos, necess�rios, em int
def converte_dtype(df_para):
    for col_para in df_para.head():
        dado_para = []
        if (col_para == 'AT' or
            col_para == 'AP' or
            col_para == 'TOTAL' or
            col_para == 'Cr�dito' or
            col_para == 'HA' or
            col_para == 'HR' or
            col_para == 'HR EAD' or
            col_para == 'CH (HR) Extensionista' or
            col_para == 'N�mero de Vagas' or
            col_para == 'Vagas Totais' or
            col_para == 'MT' or
            col_para == 'MP' or
            col_para == 'CH Docente Te�rica' or
            col_para == 'CH Docente Pr�tica' or
            col_para == 'CH Docente Total'):
            for dp in df_para[col_para]:
                dado_para.append(dp)
            for t_dp in range(len(dado_para)):
                try:
                    df_para[col_para][t_dp] = int(dado_para[t_dp])
                except:
                    continue
    return df_para

# Atentar a forma como a fun��o � escrita: Deve ser feita em ingl�s
# Insere as fun��es na tabela de aulas
def ajusta_funcoes(df_para):
    inicio = 14
    fim = 1
    for col_para in df_para.head():
        if col_para == 'Per�odo Matriz':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$K$9'
        elif col_para == 'Escola / Campus':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$I$9'
        elif col_para == 'Curso':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$D$9'
        elif col_para == 'TOTAL':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=SUM(J{i+inicio}:K{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,L{inicio}:L{i+(inicio-1)})'
        elif col_para == 'Cr�dito':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=L{i+inicio}'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,M{inicio}:M{i+(inicio-1)})'
        elif col_para == 'HA':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = F'=IF(AND(X{i+inicio}="Est�gio",Y{i+inicio}="Externo"),O{i+inicio},L{i+inicio}*20)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,N{inicio}:N{i+(inicio-1)})'
        elif col_para == 'HR':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=L{i+inicio}*15'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,O{inicio}:O{i+(inicio-1)})'
        elif col_para == 'HR EAD':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=IF(Z{i+inicio}="EAD",O{i+inicio},0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,P{inicio}:P{i+(inicio-1)})'
        elif col_para == 'CH (HR) Extensionista': # TODO: Falta a refer�ncia correta
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=#REF!'
                    # f'=SUM(H{i+inicio}:I{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = '=#REF!'
                    # f'=SUBTOTAL(109,Q{inicio}:Q{i+(inicio-1)})'
        elif col_para == 'N�mero de Vagas':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=M9'
        elif col_para == 'Vagas Totais':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=IFERROR(VLOOKUP(F{i+inicio},Lista!$AA$2:$AC$13,2,0),"-")'
        elif col_para == 'CH Docente Te�rica':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=ROUNDUP(IF(ISBLANK(AA{i+inicio}),(IFERROR((S{i+inicio}/T{i+inicio})*J{i+inicio},0)),0),0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AC{inicio}:AC{i+(inicio-1)})'
        elif col_para == 'CH Docente Pr�tica':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=ROUNDUP(IF(ISBLANK(AA{i+inicio}),(IFERROR((S{i+inicio}/U{i+inicio})*K{i+inicio},0)),0),0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AD{inicio}:AD{i+(inicio-1)})'
        elif col_para == 'CH Docente Total':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=SUM(AC{i+inicio}:AD{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AE{inicio}:AE{i+(inicio-1)})'
    return df_para

# Insere as valida��es de dados 
def validacao_dados(load_df_para):
    inicio = 14
    qtd_linhas = 100

    # Per�odo matriz
    val_periodo_matriz = DataValidation(type="list", 
        operator="equal",
        formula1='"2021/1,2021/2,2022/1,2022/2,2023/1,2023/2,2024/1,2024/2,2025/1,2025/2"', 
        allow_blank=True)

    # Per�odo
    val_periodo = DataValidation(type="list", 
        operator="equal",
        formula1='"1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16"', 
        allow_blank=True)

    # AT e AP
    val_at_ap = DataValidation(type="list", 
        operator="equal",
        formula1='"0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30"', 
        allow_blank=True)

    val_at_ap2 = DataValidation(type="whole", 
        operator="between",
        formula1=-1, 
        formula2=31,
        allow_blank=True)

    # Grupo da Disciplina
    val_g_disc = DataValidation(type="list", 
        operator="equal",
        formula1='"Curso,Escola,Institucional"',    
        allow_blank=True)

    # Tipo da Disciplina
    val_t_disc = DataValidation(type="list", 
        operator="equal",
        formula1='"Normal,Est�gio,TCC,OMC"', 
        allow_blank=True)
    
    # Modalidade
    val_modalidade = DataValidation(type="list",
        operator="equal", 
        formula1='"EAD,Presencial"', 
        allow_blank=True)
    
    # Curso Pagante
    cursos_pagantes = ['Administra��o','Administra��o - Internacional','Agronomia','An�lise De Sistemas','An�lise E Desenvolvimento De Sistemas','Arquitetura E Urbanismo','Arquitetura E Urbanismo Com �nfase Internacional','Bacharelado Em Artes Visuais','Bacharelado Em Biologia','Bacharelado Em Ci�ncias' 'Sociais','Bacharelado Em Dan�a',
        'Bacharelado Em Educa��o F�sica','Bacharelado Em Filosofia','Bacharelado Em Hist�ria','Bacharelado Em Jogos Digitais','Bacharelado Em Letras Ingl�s Internacional','Bacharelado Em Teologia','Bacharelado Interdisciplinar Em Ci�ncias E Humanidades','Bacharelado Interdisciplinar Em Sa�de','Big Data E Intelig�ncia Anal�tica','Biotecnologia','Ciberseguran�a','Ci�ncia Da Computa��o',
        'Ci�ncias Biol�gicas - Bacharelado','Ci�ncias Cont�beis','Ci�ncias Cont�beis - Internacional','Ci�ncias Da Religi�o (Licenciatura)','Ci�ncias Econ�micas','Ci�ncias Econ�micas - Internacional','Cinema E Audiovisual','Comunica��o Digital','Comunica��o Social - Hab. Publicidade E Propaganda','Comunica��o Social - Hab.: Jornalismo','Comunica��o Social - Hab.: Rela��es P�blicas',
        'Desenho Industrial','Desenho Industrial - Design De Moda','Desenho Industrial - Hab.: Design Digital','Desenho Industrial - Hab.: Programa��o Visual','Desenho Industrial - Hab.: Projeto Do Produto','Design','Design De Moda','Design De Produto','Design Digital','Design Gr�fico','Digital Data Marketing','Direito','Educa��o F�sica','Enfermagem','Engenharia','Engenharia Ambiental',
        'Engenharia Biom�dica','Engenharia Civil','Engenharia De Alimentos','Engenharia De Alimentos (�nfase Em Agroind�stria)','Engenharia De Computa��o','Engenharia De Controle E Automa��o','Engenharia De Materiais E Nanotecnologia','Engenharia De Produ��o','Engenharia De Produ��o Mec�nica','Engenharia De Redes De Comunica��es','Engenharia De Software',
        'Engenharia El�trica - Eixos: Telecomunica��es Eletr�nica Ou Sistema De Pot�ncia E Energia','Engenharia El�trica (�nfase Em Telecomunica��es)','Engenharia Eletr�nica','Engenharia Florestal','Engenharia Mec�nica','Engenharia Mecatr�nica','Engenharia Mecatr�nica (Controle E Automa��o)','Engenharia Qu�mica','Est�tica E Cosm�tica','Farm�cia','F�sica','Fisioterapia','Fonoaudiologia',
        'Gest�o Comercial','Gest�o Da Tecnologia Da Informa��o','Gest�o De Marketing Em M�dias Digitais','Gest�o De Recursos Humanos','Gest�o Financeira','Gest�o P�blica','Intelig�ncia Artificial Aplicada','Intercambio Curitiba','Interdisciplinar Em Neg�cios','International Business Program Ibp','Isoladas Curitiba','Jornalismo','Letras - Portugu�s E Ingl�s','Letras-Portugu�s-Ingl�s',
        'Licenciatura Em Artes Visuais','Licenciatura Em Ci�ncias Biol�gicas','Licenciatura Em Ci�ncias Sociais','Licenciatura Em Dan�a','Licenciatura Em Educa��o F�sica','Licenciatura Em Filosofia','Licenciatura Em F�sica','Licenciatura Em Hist�ria','Licenciatura Em Letras - Habilita��o: Ingl�s','Licenciatura Em Letras-Portugu�s','Licenciatura Em Letras-Portugu�s-Espanhol','Licenciatura Em Matem�tica',
        'Licenciatura Em M�sica','Licenciatura Em Qu�mica','Licenciatura Em Sociologia','Licenciatura Em Teatro','Log�stica','Marketing,Marketing - Internacional','Matem�tica','Medicina','Medicina Veterin�ria','Neg�cios Digitais','Nutri��o','Odontologia','Pedagogia','Processos Gerenciais','Produ��o Musical','Psicologia','Qu�mica','Rela��es P�blicas','Servi�o Social','Sistemas De Informa��o','Summer Institute',
        'Sup.De Forma��o De Prof�. P/ Ed. Infantil E S�ries Iniciais Do Ensino Fundamental','Superior De Tecnologia Em Cidades Inteligentes','Superior De Tecnologia Em Ci�ncias E Inova��o Dos Alimentos','Superior De Tecnologia Em Design De Interiores','Superior De Tecnologia Em Gastronomia','Superior De Tecnologia Em Gest�o Comercial','Superior De Tecnologia Em Gest�o Da Produ��o Industrial',
        'Superior De Tecnologia Em Gest�o Da Qualidade','Superior De Tecnologia Em Gest�o De Recursos Humanos','Superior De Tecnologia Em Gest�o De Seguran�a Privada','Superior De Tecnologia Em Gest�o Financeira','Superior De Tecnologia Em Jogos Digitais','Superior De Tecnologia Em Log�stica','Superior De Tecnologia Em Produ��o Digital Multiplataformas','Superior De Tecnologia Em Seguran�a Da Informa��o',
        'Teatro','Tecnologia Em An�lise E Desenvolvimento De Sistemas','Tecnologia Em Big Data E Intelig�ncia Anal�tica','Tecnologia Em Gest�o Comercial','Tecnologia Em Gest�o Da Tecnologia Da Informa��o','Tecnologia Em Gest�o De Marketing Em M�dias Digitais','Tecnologia Em Gest�o De Recursos Humanos','Tecnologia Em Gest�o Financeira','Tecnologia Em Gest�o Hospitalar','Tecnologia Em Log�stica',
        'Tecnologia Em Processos Gerenciais','Tradutor E Int�rprete De L�ngua Espanhola']
    
    for c in range(len(cursos_pagantes)):
        load_df_para[f'DD{c+200}'].value = cursos_pagantes[c]

    val_pagante = DataValidation(type="list", 
        operator='equal',
        formula1="$DD$200:$DD$400",
        allow_blank=True)

    # Adiconando no worksheet as valida��es
    load_df_para.add_data_validation(val_periodo_matriz)
    load_df_para.add_data_validation(val_periodo)
    load_df_para.add_data_validation(val_at_ap)
    load_df_para.add_data_validation(val_at_ap2)
    load_df_para.add_data_validation(val_g_disc)
    load_df_para.add_data_validation(val_t_disc)
    load_df_para.add_data_validation(val_modalidade)
    load_df_para.add_data_validation(val_pagante)

    val_periodo_matriz.add('K9')
    val_periodo.add(f'F{inicio}:F{qtd_linhas+(inicio-1)}')
    val_at_ap.add(f'J{inicio}:K{qtd_linhas+(inicio-1)}')
    val_at_ap2.add(f'J{inicio}:K{qtd_linhas+(inicio-1)}')
    val_g_disc.add(f'W{inicio}:W{qtd_linhas+(inicio-1)}')
    val_t_disc.add(f'X{inicio}:X{qtd_linhas+(inicio-1)}')
    val_modalidade.add(f'Z{inicio}:Z{qtd_linhas+(inicio-1)}')
    val_pagante.add(f'AA{inicio}:AA{qtd_linhas+(inicio-1)}')

    return load_df_para

# Os dados das tabelas s�o inseridos o template definido
def insere_template(load_df_para, load_template):
    fd_pagina_origem = load_df_para["Sheet1"]
    fd_pagina_destino = load_template['CRIA��O DE MATRIZ PRESENCIAL']

    # Insere os dados da tabela de aulas no modelo
    for i in range(1, fd_pagina_origem.max_row+1):
        for j in range(1, fd_pagina_destino.max_column+1):
            try:
                if(i < 101 and j<31): # Quantidade de linhas e colunas + 1 na tabela
                    fd_pagina_destino.cell(row=i+12, column=j+2).value = fd_pagina_origem.cell(row=i+12, column=j+2).value
                else:
                    break
            except:
                continue

    fd_pagina_destino = validacao_dados(fd_pagina_destino)

    ''' -------- Abre uma janela para inserir dados -------- '''
    # Curso, Escola e Total de Eletivas
    caixa_entrada = Tk()
    caixa_entrada.geometry('250x200')
    # Curso
    label_curso = Label(caixa_entrada, text="Nome do Curso")
    label_curso.pack(padx=0, pady=0, fill=X)
    nome_curso = Entry(caixa_entrada, bd =5)
    nome_curso.pack(padx=10, pady=0, fill=X)
    # Escola
    label_escola = Label(caixa_entrada, text="Nome da Escola/Campus")
    label_escola.pack(padx=0, pady=0, fill=X)
    nome_escola = Entry(caixa_entrada, bd =5)
    nome_escola.pack(padx=10, pady=0, fill=X)
    # Total de Eletivas
    label_eletiva = Label(caixa_entrada, text="Total de Eletivas")
    label_eletiva.pack(padx=0, pady=0, fill=X)
    nome_eletiva = Entry(caixa_entrada, bd =5)
    nome_eletiva.pack(padx=10, pady=0, fill=X)
    # Encerrar janela
    encerrar = Button(caixa_entrada, text ="Concluir", command = lambda:caixa_entrada.quit())
    encerrar.pack(pady = 10)
    caixa_entrada.mainloop()
    ''' -------- Encerra a janela para inserir dados -------- '''

    # Insere os dados informados na tabela
    fd_pagina_destino['D9'].value = str(nome_curso.get())
    fd_pagina_destino['I9'].value = str(nome_escola.get())
    fd_pagina_destino['D117'].value = int(nome_eletiva.get())

    # A planilha � salva
    load_template.save(f'matriz_{str(nome_curso.get())}.xlsx') 

# Fun��o principal
def main():
    # Sele��o de arquivo dos aulas
    print("Selecione a planilha de aulas")
    Tk().withdraw()
    planilha_aulas = askopenfilename(
        filetypes=[('Arquivo excel', '.xlsx')],
        title='Selecione a planilha de aulas')
    # Mostra qual foi o arquivo selecionado das aulas
    print("Arquivo selecionado: " + str(planilha_aulas).split('/')[-1])
    
    # Selecionando a planilha PARA
    print("Selecione a planilha PARA")
    Tk().withdraw()
    planilha_para = askopenfilename(
        filetypes=[('Arquivo excel', '.xlsx')],
        title='Selecione a planilha PARA')
    # Mostra qual foi o arquivo selecionado PARA
    print("Arquivo selecionado: " + str(planilha_para).split('/')[-1])

    ''' -------------------------------------------------------------------------------
        Leitura das p�ginas do arquivo
        ------------------------------------------------------------------------------- '''
    
    print("Lendo arquivo")
    
    # Planilha de aulas � armazenado na vari�vel
    df_aulas = pd.read_excel(planilha_aulas, dtype=str, header=1)

    # planilha PARA e armazenada em uma variavel, selecionando apartir do header 12
    df_para = pd.read_excel(planilha_para, dtype=str, header=12, skipfooter=5, index_col=1, sheet_name=0) # Seleciona os dados das aulas
    
    ''' -------------------------------------------------------------------------------
        Ajustes dos dados do arquivo
        ------------------------------------------------------------------------------- '''

    # Faz os ajustes necess�rios nas colunas
    df_para = ajusta_colunas(df_para)

    # Cria uma coluna para colocar o n�mero do per�odo
    df_aulas.insert(loc=0, column='Per�odo', value=nan) 
    # Ajusta as informa��es e linhas da tabela de aulas
    df_aulas = ajusta_info(df_aulas) 
    
    # remove a linha quando todas as c�lulas estiverem vazias
    df_aulas.dropna(inplace=True, how='all') 
    # remove as linhas duplicadas, nesse caso, os cabe�alhos que n�o ser�o utilizados
    df_aulas.drop_duplicates(inplace=True) 
    df_para.drop_duplicates(inplace=True) 
    # Remove os nan das c�lulas vazias
    df_aulas = df_aulas.fillna("")
    df_para = df_para.fillna("")
    
    # Insere os dados do dataframe de aulas no formato do template
    df_para = insere_dados(df_para, df_aulas) 
    # Insere as fun��es das aulas no local correto
    df_para = ajusta_funcoes(df_para) 
    # Transforma os dados necess�rios em int
    df_para = converte_dtype(df_para)

    ''' -------------------------------------------------------------------------------
        Escrita no arquivo
        ------------------------------------------------------------------------------- '''
    
    # Cria a planilha com os dados de aulas no formato do template
    df_para.to_excel('tabela_aulas_matriz.xlsx', index=False, startrow=12, startcol=2) 
    
    # Acessa a planilha com os dados
    load_df_para = load_workbook('tabela_aulas_matriz.xlsx') 
    # Acessa o template
    load_template = load_workbook(planilha_para)
    
    # Insere os dados no template e salva a planilha
    insere_template(load_df_para, load_template)

if __name__ == '__main__':
    # Chamada da fun��o main
    main()
