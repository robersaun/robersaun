# -*- coding: Windows-1252 -*-

from tkinter import Tk
from tkinter.filedialog import askopenfilename
from numpy import nan
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Ajusta as informa��es da tabela de aulas por per�odo
def ajusta_info(df_aulas):
    cont = 1 # Contador do per�odo
    linha = 0 # Contador de linha
    # Insere o n�mero do per�odo na coluna criada
    # Tira os valores desnecess�rios das linhas
    nome_coluna = [] # Cria uma lista contendo o nome de cada coluna
    for c in df_aulas.columns:
        nome_coluna.append(c)
    for dado in df_aulas['Ordem']:
        if (dado == f'{cont+1}� PER�ODO'):
            cont+=1
        if (str(dado)!='nan'):
            df_aulas['Per�odo'][linha] = str(int(cont))
            if (str(dado)=='Ordem' or str(dado)=='TOTAL' or str(dado)==f'{cont}� PER�ODO'):
                df_aulas['Per�odo'][linha] = nan
                coluna = 0
                while coluna < 11:
                    df_aulas[nome_coluna[coluna]][linha] = nan
                    coluna+=1
        linha+=1
    return df_aulas

# Verifica quais colunas possuem dados a serem validados
# Apaga as colunas desnecess�rias/n�o nomeadas
def ajusta_colunas(df):
    colunas = []
    drop_colunas = []
    contador = 0
    for c in df.head():
        # Feito desse modo para assegurar que a coluna ser� encontrada
        if (c != f'Unnamed: {contador-1}' and 
            c != f'Unnamed: {contador}' and 
            c != f'Unnamed: {contador+1}'):
            colunas.append(str(c))
        else:
            drop_colunas.append(str(c))
        contador+=1
    for dc in range(len(drop_colunas)):
        df.drop(drop_colunas[dc],inplace=True,axis=1)
    return df

# Insere os dados conforme o template
def insere_dados(df_para, df_aulas):
    for col_aulas in df_aulas.head():
        for col_para in df_para.head():
            dados_col_aula = []
            cont_dados_aulas = 0
            if (' '+col_aulas == col_para or
                col_aulas == col_para or
                col_aulas+' ' == col_para or
                (col_aulas == 'CRED' and col_para == 'Cr�dito')):
                for dados_aula in df_aulas[col_aulas]:
                    dados_col_aula.append(str(dados_aula))
                    cont_dados_aulas+=1
                for i in range(len(dados_col_aula)):
                    try:
                        # Se o dado inserido puder ser transformado em int, faz a convers�o
                        df_para[col_para][i] = int(dados_col_aula[i].split('.')[0])
                    except:
                        # Caso contr�rio, � armazenado como String
                        df_para[col_para][i] = dados_col_aula[i]
    return df_para

# Transforma os valores n�mericos, necess�rios, em int
# >> Apenas na tabela de aulas > TODO: Fazer um para a tabela de totais
# >> Antes da altera��o da forma como os dados s�o lidos, deveria ocorrer com todos os dados
def converte_dtype(df_para):
    for col_para in df_para.head():
        dado_para = []
        if (col_para == 'AT' or
            col_para == 'AP' or
            col_para == 'TOTAL' or
            col_para == 'Cr�dito' or
            col_para == 'HA' or
            col_para == 'HR' or
            col_para == 'HR EAD' or
            col_para == 'CH (HR) Extensionista' or
            col_para == 'N�mero de Vagas' or
            col_para == 'Vagas Totais' or
            col_para == 'MT' or
            col_para == 'MP' or
            col_para == 'CH Docente Te�rica' or
            col_para == 'CH Docente Pr�tica' or
            col_para == 'CH Docente Total'):
            for dp in df_para[col_para]:
                dado_para.append(dp)
            for t_dp in range(len(dado_para)):
                try:
                    df_para[col_para][t_dp] = int(dado_para[t_dp])
                except:
                    continue
    return df_para

# Atentar a forma como a fun��o � escrita: Deve ser feita em ingl�s
# Insere as fun��es na tabela de aulas
def ajusta_funcoes(df_para):
    inicio = 14
    fim = 1
    for col_para in df_para.head():
        if col_para == 'Per�odo Matriz':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$K$9'
        elif col_para == 'Escola / Campus':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$I$9'
        elif col_para == 'Curso':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$D$9'
        elif col_para == 'TOTAL':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=SUM(J{i+inicio}:K{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,L{inicio}:L{i+(inicio-1)})'
        elif col_para == 'Cr�dito':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=L{i+inicio}'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,M{inicio}:M{i+(inicio-1)})'
        elif col_para == 'HA':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = F'=IF(AND(x{i+inicio}="Est�gio",Y{i+inicio}="Externo"),O{i+inicio},L{i+inicio}*20)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,N{inicio}:N{i+(inicio-1)})'
        elif col_para == 'HR':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=L{i+inicio}*15'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,O{inicio}:O{i+(inicio-1)})'
        elif col_para == 'HR EAD':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=IF(Z{i+inicio}="EAD",M{i+inicio},0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,P{inicio}:P{i+(inicio-1)})'
        elif col_para == 'CH (HR) Extensionista': # TODO: Falta a refer�ncia correta
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=#REF!'
                    # f'=SUM(H{i+inicio}:I{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = '=#REF!'
                    # f'=SUBTOTAL(109,Q{inicio}:Q{i+(inicio-1)})'
        elif col_para == 'N�mero de Vagas':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=M9'
        elif col_para == 'Vagas Totais':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=IFERROR(VLOOKUP(F{i+inicio},Lista!$AA$2:$AC$13,2,0),"-")'
        elif col_para == 'CH Docente Te�rica':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=ROUNDUP(IF(ISBLANK(AA{i+inicio}),(IFERROR((S{i+inicio}/T{i+inicio})*J{i+inicio},0)),0),0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AC{inicio}:AC{i+(inicio-1)})'
        elif col_para == 'CH Docente Pr�tica':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=ROUNDUP(IF(ISBLANK(AA{i+inicio}),(IFERROR((S{i+inicio}/U{i+inicio})*K{i+inicio},0)),0),0)'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AD{inicio}:AD{i+(inicio-1)})'
        elif col_para == 'CH Docente Total':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = f'=SUM(AC{i+inicio}:AD{i+inicio})'
                if ((((len(df_para[col_para])-i) - fim)) == 0):
                    df_para[col_para][i] = f'=SUBTOTAL(109,AE{inicio}:AE{i+(inicio-1)})'
    return df_para

# TODO: Inserir as fun��es da tabela de totais >> N�o est� funcionando
def ajusta_funcoes_t2(df_para):
    inicio = 117
    for col_para in df_para.head():
        if col_para == 'CR�DITO':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) == 0):
                    df_para[col_para][i] = f'=D{inicio}'
        """elif col_para == 'HORA AULA':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$K$9'
        elif col_para == 'HORA REL�GIO':
            for i in range(len(df_para[col_para])):
                if ((len(df_para[col_para])-i) - fim > 0):
                    df_para[col_para][i] = '=$K$9'"""
    return df_para

def validacao_dados(load_df_para):
    inicio = 14
    qtd_linhas = 100

    # Per�odo matriz
    val_periodo_matriz = DataValidation(type="list", 
        operator="equal",
        formula1='"2021/1,2021/2,2022/1,2022/2,2023/1,2023/2,2024/1,2024/2,2025/1,2025/2"', 
        allow_blank=True)

    # Per�odo
    val_periodo = DataValidation(type="list", 
        operator="equal",
        formula1='"1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16"', 
        allow_blank=True)

    # AT e AP
    val_at_ap = DataValidation(type="list", 
        operator="equal",
        formula1='"0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30"', 
        allow_blank=True)

    val_at_ap2 = DataValidation(type="whole", 
        operator="between",
        formula1=-1, 
        formula2=31,
        allow_blank=True)

    # Grupo da Disciplina
    val_g_disc = DataValidation(type="list", 
        operator="equal",
        formula1='"Curso,Escola,Institucional"',    
        allow_blank=True)

    # Tipo da Disciplina
    val_t_disc = DataValidation(type="list", 
        operator="equal",
        formula1='"Normal,Est�gio,TCC,OMC"', 
        allow_blank=True)
    
    # Modalidade
    val_modalidade = DataValidation(type="list",
        operator="equal", 
        formula1='"EAD,Presencial"', 
        allow_blank=True)
    
    # Curso Pagante #TODO: Colocar aquela lista de cursos q tem no template
    # Curso Pagante >> N�o funciona e tira a valida��o do resto da planilha >> Talvez pegar de outra parte da planilha?
    # val_pagante = DataValidation(type="list", 
    #    formula1=('"Administra��o,Administra��o - Internacional,Agronomia,An�lise De Sistemas,An�lise E Desenvolvimento De Sistemas,Arquitetura E Urbanismo,Arquitetura E Urbanismo Com �nfase Internacional,Bacharelado Em Artes Visuais,Bacharelado Em Biologia,Bacharelado Em Ci�ncias Sociais,Bacharelado Em Dan�a,'
    #    'Bacharelado Em Educa��o F�sica,Bacharelado Em Filosofia,Bacharelado Em Hist�ria,Bacharelado Em Jogos Digitais,Bacharelado Em Letras Ingl�s Internacional,Bacharelado Em Teologia,Bacharelado Interdisciplinar Em Ci�ncias E Humanidades,Bacharelado Interdisciplinar Em Sa�de,Big Data E Intelig�ncia Anal�tica,Biotecnologia,Ciberseguran�a,Ci�ncia Da Computa��o,'
    #    'Ci�ncias Biol�gicas - Bacharelado,Ci�ncias Cont�beis,Ci�ncias Cont�beis - Internacional,Ci�ncias Da Religi�o (Licenciatura),Ci�ncias Econ�micas,Ci�ncias Econ�micas - Internacional,Cinema E Audiovisual,Comunica��o Digital,Comunica��o Social - Hab. Publicidade E Propaganda,Comunica��o Social - Hab.: Jornalismo,Comunica��o Social - Hab.: Rela��es P�blicas,'
    #    'Desenho Industrial,Desenho Industrial - Design De Moda,Desenho Industrial - Hab.: Design Digital,Desenho Industrial - Hab.: Programa��o Visual,Desenho Industrial - Hab.: Projeto Do Produto,Design,Design De Moda,Design De Produto,Design Digital,Design Gr�fico,Digital Data Marketing,Direito,Educa��o F�sica,Enfermagem,Engenharia,Engenharia Ambiental,'
    #    'Engenharia Biom�dica,Engenharia Civil,Engenharia De Alimentos,Engenharia De Alimentos (�nfase Em Agroind�stria),Engenharia De Computa��o,Engenharia De Controle E Automa��o,Engenharia De Materiais E Nanotecnologia,Engenharia De Produ��o,Engenharia De Produ��o Mec�nica,Engenharia De Redes De Comunica��es,Engenharia De Software,'
    #    'Engenharia El�trica - Eixos: Telecomunica��es Eletr�nica Ou Sistema De Pot�ncia E Energia,Engenharia El�trica (�nfase Em Telecomunica��es),Engenharia Eletr�nica,Engenharia Florestal,Engenharia Mec�nica,Engenharia Mecatr�nica,Engenharia Mecatr�nica (Controle E Automa��o),Engenharia Qu�mica,Est�tica E Cosm�tica,Farm�cia,F�sica,Fisioterapia,Fonoaudiologia,'
    #    'Gest�o Comercial,Gest�o Da Tecnologia Da Informa��o,Gest�o De Marketing Em M�dias Digitais,Gest�o De Recursos Humanos,Gest�o Financeira,Gest�o P�blica,Intelig�ncia Artificial Aplicada,Intercambio Curitiba,Interdisciplinar Em Neg�cios,International Business Program Ibp,Isoladas Curitiba,Jornalismo,Letras - Portugu�s E Ingl�s,Letras-Portugu�s-Ingl�s,'
    #    'Licenciatura Em Artes Visuais,Licenciatura Em Ci�ncias Biol�gicas,Licenciatura Em Ci�ncias Sociais,Licenciatura Em Dan�a,Licenciatura Em Educa��o F�sica,Licenciatura Em Filosofia,Licenciatura Em F�sica,Licenciatura Em Hist�ria,Licenciatura Em Letras - Habilita��o: Ingl�s,Licenciatura Em Letras-Portugu�s,Licenciatura Em Letras-Portugu�s-Espanhol,Licenciatura Em Matem�tica,'
    #    'Licenciatura Em M�sica,Licenciatura Em Qu�mica,Licenciatura Em Sociologia,Licenciatura Em Teatro,Log�stica,Marketing,Marketing - Internacional,Matem�tica,Medicina,Medicina Veterin�ria,Neg�cios Digitais,Nutri��o,Odontologia,Pedagogia,Processos Gerenciais,Produ��o Musical,Psicologia,Qu�mica,Rela��es P�blicas,Servi�o Social,Sistemas De Informa��o,Summer Institute,'
    #    'Sup.De Forma��o De Prof�. P/ Ed. Infantil E S�ries Iniciais Do Ensino Fundamental,Superior De Tecnologia Em Cidades Inteligentes,Superior De Tecnologia Em Ci�ncias E Inova��o Dos Alimentos,Superior De Tecnologia Em Design De Interiores,Superior De Tecnologia Em Gastronomia,Superior De Tecnologia Em Gest�o Comercial,Superior De Tecnologia Em Gest�o Da Produ��o Industrial,'
    #    'Superior De Tecnologia Em Gest�o Da Qualidade,uperior De Tecnologia Em Gest�o De Recursos Humanos,Superior De Tecnologia Em Gest�o De Seguran�a Privada,Superior De Tecnologia Em Gest�o Financeira,Superior De Tecnologia Em Jogos Digitais,Superior De Tecnologia Em Log�stica,Superior De Tecnologia Em Produ��o Digital Multiplataformas,Superior De Tecnologia Em Seguran�a Da Informa��o,'
    #    'Teatro,Tecnologia Em An�lise E Desenvolvimento De Sistemas,Tecnologia Em Big Data E Intelig�ncia Anal�tica,Tecnologia Em Gest�o Comercial,Tecnologia Em Gest�o Da Tecnologia Da Informa��o,Tecnologia Em Gest�o De Marketing Em M�dias Digitais,Tecnologia Em Gest�o De Recursos Humanos,Tecnologia Em Gest�o Financeira,Tecnologia Em Gest�o Hospitalar,Tecnologia Em Log�stica,'
    #    'Tecnologia Em Processos Gerenciais,Tradutor E Int�rprete De L�ngua Espanhola"'), 
    #    allow_blank=True)

    # Adiconando no worksheet as valida��es
    load_df_para.add_data_validation(val_periodo_matriz)
    load_df_para.add_data_validation(val_periodo)
    load_df_para.add_data_validation(val_at_ap)
    load_df_para.add_data_validation(val_at_ap2)
    load_df_para.add_data_validation(val_g_disc)
    load_df_para.add_data_validation(val_t_disc)
    load_df_para.add_data_validation(val_modalidade)
    #load_df_para.add_data_validation(val_pagante)

    val_periodo_matriz.add('K9')
    val_periodo.add(f'F{inicio}:F{qtd_linhas+(inicio-1)}')
    val_at_ap.add(f'J{inicio}:K{qtd_linhas+(inicio-1)}')
    val_at_ap2.add(f'J{inicio}:K{qtd_linhas+(inicio-1)}')
    val_g_disc.add(f'W{inicio}:W{qtd_linhas+(inicio-1)}')
    val_t_disc.add(f'X{inicio}:X{qtd_linhas+(inicio-1)}')
    val_modalidade.add(f'Z{inicio}:Z{qtd_linhas+(inicio-1)}')
    #val_pagante.add(f'AA{inicio}:AA{qtd_linhas+(inicio-1)}')

    return load_df_para

# Os dados das tabelas s�o inseridos o template definido
def insere_template(load_df_para, load_df_para_2, load_template):
    fd_pagina_origem = load_df_para["Sheet1"]
    fd_pagina_origem_totais = load_df_para_2["Sheet1"]
    fd_pagina_destino = load_template['CRIA��O DE MATRIZ PRESENCIAL']

    # Insere os dados da tabela de aulas no modelo
    for i in range(1, fd_pagina_origem.max_row+1):
        for j in range(1, fd_pagina_destino.max_column+1):
            try:
                if(i < 101):
                    fd_pagina_destino.cell(row=i+12, column=j+2).value = fd_pagina_origem.cell(row=i+12, column=j+2).value
                else:
                    break
            except:
                continue

    # Insere os dados da tabela de totais no modelo
    for i in range(1, fd_pagina_origem.max_row+1):
        for j in range(1, fd_pagina_destino.max_column+1):
            try:
                fd_pagina_destino.cell(row=i+115, column=j+3).value = fd_pagina_origem_totais.cell(row=i+115, column=j+3).value
            except:
                continue

    fd_pagina_destino = validacao_dados(fd_pagina_destino)

    load_template.save('matriz_template.xlsx') # A matriz inserida no template

# Fun��o principal
def main():
    # Sele��o de arquivo dos aulas
    print("Selecione a planilha de aulas")
    Tk().withdraw()
    planilha_aulas = askopenfilename(
        filetypes=[('Arquivo excel', '.xlsx')],
        title='Selecione a planilha de aulas')
    # Mostra qual foi o arquivo selecionado das aulas
    print("Arquivo selecionado: " + str(planilha_aulas).split('/')[-1])
    
    # Selecionando a planilha PARA
    print("Selecione a planilha PARA")
    Tk().withdraw()
    planilha_para = askopenfilename(
        filetypes=[('Arquivo excel', '.xlsx')],
        title='Selecione a planilha PARA')
    # Mostra qual foi o arquivo selecionado PARA
    print("Arquivo selecionado: " + str(planilha_para).split('/')[-1])

    ''' -------------------------------------------------------------------------------
        Leitura das p�ginas do arquivo
        ------------------------------------------------------------------------------- '''
    
    print("Lendo arquivo")
    
    # Planilha de aulas � armazenado na vari�vel
    df_aulas = pd.read_excel(planilha_aulas, dtype=str, header=1)

    # planilha PARA e armazenada em uma variavel, selecionando apartir do header 12
    df_para = pd.read_excel(planilha_para, dtype=str, header=12, skipfooter=5, index_col=1, sheet_name=0) # Seleciona os dados das aulas
    df_para2 = pd.read_excel(planilha_para, header=115, index_col=1, sheet_name=0) # Identifica os dados dos totais inidicados

    df_para = ajusta_colunas(df_para) # Faz os ajustes necess�rios nas colunas
    df_para2 = ajusta_colunas(df_para2)

    df_aulas.insert(loc=0, column='Per�odo', value=nan) # Cria uma coluna para colocar o n�mero do per�odo
    df_aulas = ajusta_info(df_aulas) # Ajusta as informa��es e linhas da tabela de aulas
    
    # Remove os nan das c�lulas vazias
    df_aulas.dropna(inplace=True, how='all') # remove a linha, quando todas as c�lulas estiverem vazias
    df_aulas.drop_duplicates(inplace=True) # remove duplicadas, no caso, os espa�os em branco e os t�tulos
    df_aulas = df_aulas.fillna("")
    
    df_para.drop_duplicates(inplace=True) # remove duplicadas, nesse caso, os espa�os em branco e os t�tulos
    df_para = df_para.fillna("")
    
    df_para = insere_dados(df_para, df_aulas) # Pega os dados do dataframe de alunos e insere no template
    df_para = ajusta_funcoes(df_para) # Insere as fun��es da tabela 1 no local correto
    df_para2 = ajusta_funcoes_t2(df_para2)
    df_para = converte_dtype(df_para)

    '''------------------------------------------------------------------------------------------------'''
    
    df_para.to_excel('tabela_aulas_matriz.xlsx', index=False, startrow=12, startcol=2) # Cria a planilha utilizando o template PARA
    df_para2.to_excel('tabela_totais_matriz.xlsx', index=False, startrow=115, startcol=3) # Isso era para mexer na tabela 2

    load_df_para = load_workbook('tabela_aulas_matriz.xlsx') # Cria o arquivo com o nome do curso
    load_df_para_2 = load_workbook('tabela_totais_matriz.xlsx') # Cria o arquivo com os totais
    load_template = load_workbook(planilha_para)
    
    insere_template(load_df_para, load_df_para_2, load_template)

if __name__ == '__main__':
    # Chamada da fun��o main
    main()
