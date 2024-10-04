# -*- coding: Windows-1252 -*-

from tkinter import Tk, filedialog
import xlrd as x
from openpyxl import Workbook as wb
import pyexcel as p
import pandas as pd
import os


# Função principal
# noinspection PyBroadException
from xlrd.timemachine import xrange


def main(root):
    # Para cada diretório dentro do diretório raiz
    for dirpath, dirnames, planilha in os.walk(root):

        # Se o diretório possuir outros diretórios dentro dele, ignora
        if dirnames:
            continue
        # Seleção do arquivo

        curso = dirpath.replace('\\', '/').split('/')[-1]

        # nome_arquivo = str(planilha).split('/')[-1]
        # Mostra qual foi o arquivo selecionado
        for nome_arquivo in planilha:
            #try:
                print(f"Arquivo selecionado: {nome_arquivo}")
                doc_nome = nome_arquivo.split('.')[0]
                doc_grade = f"{nome_arquivo.split('_')[-2]}-{nome_arquivo.split('_')[-1].split('.')[0]}"
                os.makedirs(f'{curso}/{doc_grade}', exist_ok=True)

                ''' -------------------------------------------------------------------------------
                    Leitura das páginas do arquivo
                    ------------------------------------------------------------------------------- '''

                # Planilha de equivalências é armazenada na variável
                # Os dados dos alunos deverão sempre estar na primeira página do arquivo
                arquivo = f'{dirpath}/{nome_arquivo}'
                df_f = x.open_workbook(arquivo)
                workbook = wb()

                for i in xrange(0, df_f.nsheets):
                    df_s = df_f.sheet_by_index(i)
                    sheet = workbook.active if i == 0 else workbook.create_sheet()
                    sheet.title = df_s.name

                    for row in df_f(0, df_s.nrows):
                        for col in df_f(0, df_s.ncols):
                            sheet.cell(row=row + 1, column=col + 1).value = df_s.cell_value(row, col)

                """sheet_names = df_f.sheetnames()
                dfn = []
                for s in sheet_names:
                    dfn.append(pd.read_excel(arquivo, dtype=str, sheet_name=s))
                print(sheet_names)
                with pd.ExcelWriter(f'{curso}/{doc_grade}/{doc_nome}.xlsx') as writer:
                    for c in range(len(sheet_names)):
                        sn = sheet_names[c]
                        header = []
                        for h in dfn[c].head():
                            header.append('')
                        if 'Grup ' in sn:
                            sn = 'Grup Atv Comp - Proj Com'
                        dfn[c].to_excel(writer, index=False, header=False, sheet_name=sn, startrow=1)
                print('Excel gerado')"""
            #except:
            #    pass


if __name__ == '__main__':
    # Chamada da função main
    Tk().withdraw()
    pasta_raiz = filedialog.askdirectory()
    main(pasta_raiz)

# Aparentemnte agr é só selecionar o diretório raiz q roda td
