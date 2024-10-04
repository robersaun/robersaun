import os
from tkinter import filedialog, Tk

import openpyxl
import pandas


def main():
    Tk().withdraw()
    diretorio = filedialog.askdirectory()
    escola = diretorio.split('/')[-1]

    count = 0
    duplicado = 0

    disciplinas = []
    for dirpath, dirnames, filenames in os.walk(diretorio):
        for file in filenames:
            if "xlsx" in file and 'PUC' not in file:
                count += 1
                caminho_completo = dirpath + "/" + file
                caminho_completo = caminho_completo.replace("\\", "/")

                print(f"Lendo arquivo {count}: {caminho_completo.replace(diretorio, '')}")
                try:
                    excel = ler_excel(caminho_completo)
                    for disciplina in excel:
                        disciplinas.append(disciplina)
                except Exception as e:
                    print(f"\tNão foi possível ler o arquivo, erro: {e}")
            if 'PUC' in file:
                duplicado += 1
                print(f'{duplicado} arquivo(s) duplicado(s) encontrado(s)')
    print(f"Total de arquivos: {count}")

    dataframe = pandas.DataFrame(disciplinas)
    dataframe.columns = ['Disciplina', 'Período', 'Classificação', 'Créditos', 'Grupo', 'CH Teórica',
                         'CH Prática', 'CH Oficial', 'CH Relógio Oficial', 'Curso', 'Matriz', 'Equivalência(s)']
    dataframe.to_excel(excel_writer=f"Unificado com Ciclos/Excel-Unificado-{escola}-v2.xlsx", index=False)


def ler_excel(arquivo):
    disciplinas_excel = []
    wb = openpyxl.load_workbook(arquivo)
    sheet_excel = wb["Grade Curricular"]
    matriz = str(sheet_excel.cell(2, 3).value)
    curso = str(sheet_excel.cell(3, 3).value)

    sheet_excel = wb["Ciclos"]
    periodo = 0

    for linha in range(1, sheet_excel.max_row):
        if str(sheet_excel.cell(linha, 2).value) == "Ciclo:":
            periodo = int(sheet_excel.cell(linha, 3).value)

        if str(sheet_excel.cell(linha, 2).value).replace('None', '') == "" or \
                str(sheet_excel.cell(linha, 2).value) == "Total Carga Horária (mínimo):" or \
                str(sheet_excel.cell(linha, 2).value) == "Disciplina" or \
                str(sheet_excel.cell(linha, 2).value) == "Ciclo:" or \
                str(sheet_excel.cell(linha, 2).value) == "Total":
            continue

        nome_disciplina = str(sheet_excel.cell(linha, 2).value).strip()
        classificacao = str(sheet_excel.cell(linha, 3).value)
        grupo = str(sheet_excel.cell(linha, 6).value).strip()

        ch_teorica = str(sheet_excel.cell(linha, 7).value).replace('None', '')
        if ch_teorica == "":
            ch_teorica = 0
        ch_teorica = int(ch_teorica)

        ch_pratica = str(sheet_excel.cell(linha, 8).value).replace('None', '')
        if ch_pratica == "":
            ch_pratica = 0
        ch_pratica = int(ch_pratica)

        ch_oficial = str(sheet_excel.cell(linha, 9).value).replace('None', '')
        if ch_oficial == "":
            ch_oficial = 0
        ch_oficial = int(ch_oficial)

        ch_relogio_oficial = str(sheet_excel.cell(linha, 13).value).replace('None', '')
        if ch_relogio_oficial == "":
            ch_relogio_oficial = 0
        ch_relogio_oficial = int(ch_relogio_oficial)

        creditos = str(sheet_excel.cell(linha, 16).value).replace('None', '')
        if creditos == "":
            creditos = 0
        creditos = int(creditos)

        equivalencia = str(sheet_excel.cell(linha, 19).value).replace("<br />", "\n").replace('None', '')

        disciplinas_excel.append([nome_disciplina, periodo, classificacao, creditos, grupo, ch_teorica, ch_pratica,
                                  ch_oficial, ch_relogio_oficial, curso, matriz, equivalencia])

    return disciplinas_excel


if __name__ == '__main__':
    main()
