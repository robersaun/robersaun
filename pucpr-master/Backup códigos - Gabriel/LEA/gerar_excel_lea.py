import os
from tkinter import filedialog, Tk

import openpyxl
import pandas


def main():
    Tk().withdraw()
    diretorio = filedialog.askdirectory()
    escola = diretorio.split('/')[-1]

    count = 0

    disciplinas = []
    for dirpath, dirnames, filenames in os.walk(diretorio):
        for file in filenames:
            if "xlsx" in file:
                if "Equivalencias-" in file or "Excel-Unificado-" in file:
                    continue

                count += 1
                caminho_completo = dirpath + "/" + file
                caminho_completo = caminho_completo.replace("\\", "/")

                print(f"Lendo arquivo {count}: {caminho_completo.replace(diretorio, '')}")
                try:
                    excel = ler_excel(caminho_completo)
                    for disciplina in excel:
                        disciplinas.append(disciplina)
                except Exception as e:
                    print(f"\tNão foi possível ler o arquivo, erro: {e}")
    print(f"Total de arquivos: {count}")

    dataframe = pandas.DataFrame(disciplinas)
    dataframe.columns = ['Disciplina', 'Matriz', 'Ciclo', 'Curso', 'Classificação', 'Créditos', 'Grupo', 'CH Teórica',
                         'CH Prática', 'CH Oficial', 'CH Relógio Oficial', 'Equivalência(s)']
    dataframe.to_excel(excel_writer=f"Excel-LEA-{escola}.xlsx", index=False)


def ler_excel(arquivo):
    disciplinas_excel = []
    wb = openpyxl.load_workbook(arquivo)
    sheet_excel = wb["Grade Curricular"]
    matriz = str(sheet_excel.cell(2, 3).value)
    curso = str(sheet_excel.cell(3, 3).value)
    ciclo = 0

    sheet_excel = wb["Ciclos"]
    for linha in range(1, sheet_excel.max_row):
        if str(sheet_excel.cell(linha, 2).value) == "Ciclo:":
            ciclo = int(sheet_excel.cell(linha, 3).value)
            continue

        if str(sheet_excel.cell(linha, 2).value).replace('None', '') == "" or \
                str(sheet_excel.cell(linha, 2).value) == "Total Carga Horária (mínimo):" or \
                str(sheet_excel.cell(linha, 2).value) == "Disciplina" or \
                str(sheet_excel.cell(linha, 2).value) == "Total" or \
                "EH03054-04A - Leitura e Escrita de Textos Técnico-Científicos" not in \
                str(sheet_excel.cell(linha, 2).value) and \
                "ELE100A - Leitura e Escrita Acadêmica" not in \
                str(sheet_excel.cell(linha, 2).value) and \
                "EH03051-04 - Leitura e Produção de Textos Acadêmicos" not in \
                str(sheet_excel.cell(linha, 2).value) and \
                "EH03026-04A - Leitura e Produção de Textos Acad.s" not in \
                str(sheet_excel.cell(linha, 2).value) and \
                "EH04116-04A - Leitura e Escrita Acadêmica" not in \
                str(sheet_excel.cell(linha, 2).value):
            continue

        nome_disciplina = str(sheet_excel.cell(linha, 2).value).strip()
        classificacao = str(sheet_excel.cell(linha, 3).value)
        grupo = str(sheet_excel.cell(linha, 6).value).strip()

        ch_teorica = str(sheet_excel.cell(linha, 7).value).replace('None', '')
        if ch_teorica == "":
            ch_teorica = 0
        ch_teorica = int(ch_teorica)

        ch_pratica = str(sheet_excel.cell(linha, 8).value).replace('None', '')
        if ch_pratica == "":
            ch_pratica = 0
        ch_pratica = int(ch_pratica)

        ch_oficial = str(sheet_excel.cell(linha, 9).value).replace('None', '')
        if ch_oficial == "":
            ch_oficial = 0
        ch_oficial = int(ch_oficial)

        ch_relogio_oficial = str(sheet_excel.cell(linha, 13).value).replace('None', '')
        if ch_relogio_oficial == "":
            ch_relogio_oficial = 0
        ch_relogio_oficial = int(ch_relogio_oficial)

        creditos = str(sheet_excel.cell(linha, 16).value).replace('None', '')
        if creditos == "":
            creditos = 0
        creditos = int(creditos)

        equivalencia = str(sheet_excel.cell(linha, 19).value).replace("<br />", "\n").replace('None', '')

        disciplinas_excel.append(
            [nome_disciplina, matriz, ciclo, curso, classificacao, creditos, grupo, ch_teorica, ch_pratica, ch_oficial,
             ch_relogio_oficial, equivalencia])

    return disciplinas_excel


if __name__ == '__main__':
    main()
