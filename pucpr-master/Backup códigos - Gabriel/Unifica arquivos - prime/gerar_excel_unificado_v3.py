import os
from tkinter import filedialog, Tk

import openpyxl
import pandas


def main():
    Tk().withdraw()
    diretorio = filedialog.askdirectory()
    campi = diretorio.split('/')[-1]

    count = 0
    duplicado = 0

    disciplinas = []
    for dirpath, dirnames, filenames in os.walk(diretorio):
        for file in filenames:
            if "xlsx" in file and 'PUC' not in file:
                count += 1
                caminho_completo = dirpath + "/" + file
                caminho_completo = caminho_completo.replace("\\", "/")

                print(f"Lendo arquivo {count}: {caminho_completo.replace(diretorio, '')}")
                try:
                    excel = ler_excel(caminho_completo)
                    for disciplina in excel:
                        disciplinas.append(disciplina)
                except Exception as e:
                    print(f"\tNão foi possível ler o arquivo, erro: {e}")
            if 'PUC' in file:
                duplicado += 1
                print(f'{duplicado} arquivo(s) duplicado(s) encontrado(s)')
    print(f"Total de arquivos: {count}")

    dataframe = pandas.DataFrame(disciplinas)
    dataframe.columns = ['Curso', 'Matriz',
                         'Disciplina', 'Descrição na Grade', 'Abreviação na Grade',
                         'Período', 'Classificação', 'Tipo',
                         'Créditos', 'Créditos Cobrados', 'CH Contrato', 'Grupo',
                         'CH Teórica', 'CH Prática', 'CH Oficial',
                         'CH Relógio Teórica', 'CH Relógio Prática', 'CH Relógio Oficial',
                         'CH Tutoria', 'CH Relógio Tutoria',
                         'CH Extensão', 'CH Relógio Extensão',
                         'Requisito(s)', 'Equivalência(s)']
    dataframe.to_excel(excel_writer=f"Unificados por Campus/Excel-Unificado-{campi}-v3.xlsx", index=False)


def ler_excel(arquivo):
    disciplinas_excel = []
    wb = openpyxl.load_workbook(arquivo)
    sheet_excel = wb["Grade Curricular"]
    matriz = str(sheet_excel.cell(2, 3).value)
    curso = str(sheet_excel.cell(3, 3).value)

    sheet_excel = wb["Ciclos"]
    periodo = 0

    for linha in range(1, sheet_excel.max_row):
        if str(sheet_excel.cell(linha, 2).value) == "Ciclo:":
            periodo = int(sheet_excel.cell(linha, 3).value)

        if str(sheet_excel.cell(linha, 2).value).replace('None', '') == "" or \
                str(sheet_excel.cell(linha, 2).value) == "Total Carga Horária (mínimo):" or \
                str(sheet_excel.cell(linha, 2).value) == "Disciplina" or \
                str(sheet_excel.cell(linha, 2).value) == "Ciclo:" or \
                str(sheet_excel.cell(linha, 2).value) == "Total":
            continue

        nome_disciplina = str(sheet_excel.cell(linha, 2).value).strip()
        descricao_disc = str(sheet_excel.cell(linha, 4).value).strip()
        abreviacao_disc = str(sheet_excel.cell(linha, 5).value).strip()
        classificacao = str(sheet_excel.cell(linha, 3).value)
        tipo = str(sheet_excel.cell(linha, 23).value)
        grupo = str(sheet_excel.cell(linha, 6).value).strip()

        ch_teorica = str(sheet_excel.cell(linha, 7).value).replace('None', '')
        if ch_teorica == "":
            ch_teorica = 0
        ch_teorica = int(ch_teorica)

        ch_pratica = str(sheet_excel.cell(linha, 8).value).replace('None', '')
        if ch_pratica == "":
            ch_pratica = 0
        ch_pratica = int(ch_pratica)

        ch_oficial = str(sheet_excel.cell(linha, 9).value).replace('None', '')
        if ch_oficial == "":
            ch_oficial = 0
        ch_oficial = int(ch_oficial)

        ch_contrato = str(sheet_excel.cell(linha, 10).value).replace('None', '')
        if ch_contrato == "":
            ch_contrato = 0
        ch_contrato = int(ch_contrato)

        ch_relogio_teorica = str(sheet_excel.cell(linha, 11).value).replace('None', '')
        if ch_relogio_teorica == "":
            ch_relogio_teorica = 0
        ch_relogio_teorica = int(ch_relogio_teorica)

        ch_relogio_pratica = str(sheet_excel.cell(linha, 12).value).replace('None', '')
        if ch_relogio_pratica == "":
            ch_relogio_pratica = 0
        ch_relogio_pratica = int(ch_relogio_pratica)

        ch_relogio_oficial = str(sheet_excel.cell(linha, 13).value).replace('None', '')
        if ch_relogio_oficial == "":
            ch_relogio_oficial = 0
        ch_relogio_oficial = int(ch_relogio_oficial)

        ch_tutoria = str(sheet_excel.cell(linha, 14).value).replace('None', '')
        if ch_tutoria == "":
            ch_tutoria = 0
        ch_tutoria = int(ch_tutoria)

        ch_relogio_tutoria = str(sheet_excel.cell(linha, 15).value).replace('None', '')
        if ch_relogio_tutoria == "":
            ch_relogio_tutoria = 0
        ch_relogio_tutoria = int(ch_relogio_tutoria)

        ch_extensao = str(sheet_excel.cell(linha, 16).value).replace('None', '')
        if ch_extensao == "":
            ch_extensao = 0
        ch_extensao = int(ch_extensao)

        ch_extensao_tutoria = str(sheet_excel.cell(linha, 17).value).replace('None', '')
        if ch_extensao_tutoria == "":
            ch_extensao_tutoria = 0
        ch_extensao_tutoria = int(ch_extensao_tutoria)

        creditos = str(sheet_excel.cell(linha, 18).value).replace('None', '')
        if creditos == "":
            creditos = 0
        creditos = int(creditos)

        cr_cobrados = str(sheet_excel.cell(linha, 19).value).replace('None', '')
        if cr_cobrados == "":
            cr_cobrados = 0
        cr_cobrados = int(cr_cobrados)

        requisitos = str(sheet_excel.cell(linha, 20).value).replace("<br />", "\n").replace('None', '')
        equivalencia = str(sheet_excel.cell(linha, 21).value).replace("<br />", "\n").replace('None', '')

        disciplinas_excel.append([curso, matriz,
                                  nome_disciplina, descricao_disc, abreviacao_disc,
                                  periodo, classificacao, tipo,
                                  creditos, cr_cobrados, ch_contrato, grupo,
                                  ch_teorica, ch_pratica, ch_oficial,
                                  ch_relogio_teorica, ch_relogio_pratica, ch_relogio_oficial,
                                  ch_tutoria, ch_relogio_tutoria,
                                  ch_extensao, ch_extensao_tutoria,
                                  requisitos, equivalencia])
    return disciplinas_excel


if __name__ == '__main__':
    main()

"""
2 Disciplina - ok
3 Classif. - ok
4 Descrição na Grade - ok
5 Abreviação na Grade - ok
6 Grupo - ok
7 C.H. Teórica - ok
8 C.H. Prática - ok
9 C.H. Oficial - ok
10 C.H. Contrato - ok
11 C.H. Relógio Teórica - ok
12 C.H. Relógio Prática - ok
13 C.H. Relógio Oficial - ok
14 C.H. Tutoria - ok
15 C.H. Relógio Tutoria - ok
16 C.H. Extensão - ok
17 C.H. Relógio Extensão - ok
18 Nº de Créditos - ok
19 Nº de Créditos Cobrados	- ok
20 Requisito(s) - ok
21 Equivalência(s) - ok
22 CicloMínimo - ñ
23 Tipo - ok
"""