"""Comparador de Matrizes Modulares

Script para comparar uma única matriz baixada do Prime com uma resolução no formato de módulos ao invés de semestres.
Recebe como entrada ambos os arquivos em formato Excel, e gera um relatório de diferenças também em formato Excel.
As resoluções originalmente vêm em formato Word, e para converter para Excel deve-se copiar todos os dados e colar em
uma nova planilha, cuidando para que cada disciplina fique em uma única linha.
As matrizes do Prime originalmente vêm em formato XLS, para converter em XLSX deve-se abrí-las no Excel e 'Salvar como'.

Versão descontinuada, porém script para substituição não foi desenvolvido.


Desenvolvido por Vinicius Tozo
Última atualização: 13/08/2021
"""
import xlrd
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import re

alerta = False


def tratar_requisitos(req):
    global alerta
    resultados = []
    requisitos = re.split("Pré-R|Co-Re|Requisito de Posição: |Requisito Espe", req)

    for requisito in requisitos:
        requisito = requisito.replace(",", "").strip()
        if "equisito: " in requisito:  # Pré-requisito
            disciplina = requisito.strip().split(" - ")[1:]
            separador = " – "
            disciplina = separador.join(disciplina).strip().replace("–", "-")
            existe = any(disciplina in sublist for sublist in disciplinas_word)

            if existe:
                x = [x for x in disciplinas_word if disciplina in x][0]
                ordem = disciplinas_word.index(x) + 1
            else:
                ordem = "???"
            resultados.append(str(ordem) + "PR")
        elif "quisito Direcional: " in requisito:  # Co-requisito Direcional
            disciplina = requisito.strip().split(" - ")[1:]
            separador = " – "
            disciplina = separador.join(disciplina).strip().replace("–", "-")
            existe = any(disciplina in sublist for sublist in disciplinas_word)

            if existe:
                x = [x for x in disciplinas_word if disciplina in x][0]
                ordem = disciplinas_word.index(x) + 1
            else:
                ordem = "???"
            resultados.append(str(ordem) + "CD")
        elif "quisito: " in requisito:  # Co-requisito
            disciplina = requisito.strip().split(" - ")[1:]
            separador = " – "
            disciplina = separador.join(disciplina).strip().replace("–", "-")
            existe = any(disciplina in sublist for sublist in disciplinas_word)

            if existe:
                x = [x for x in disciplinas_word if disciplina in x][0]
                ordem = disciplinas_word.index(x) + 1
            else:
                ordem = "???"
            resultados.append(str(ordem) + "CR")
        elif "cial: " in requisito:  # Requisito especial
            disciplina = requisito.strip().split(" - ")[1:]
            separador = " – "
            disciplina = separador.join(disciplina).strip().replace("–", "-")
            existe = any(disciplina in sublist for sublist in disciplinas_word)

            if existe:
                x = [x for x in disciplinas_word if disciplina in x][0]
                ordem = disciplinas_word.index(x) + 1
            else:
                ordem = "???"
            resultados.append(str(ordem) + "RE")
        elif "Percentual Mínimo" in requisito:
            percentual = requisito.split("= ")[-1]
            if percentual != "0":
                resultados.append("RP" + str(percentual) + "%")
        elif "Número Mínimo de Créditos Cursados" in requisito:
            n = requisito.split("= ")[-1]
            if n != "0":
                resultados.append("RP" + str(n) + "ct")
        elif "Total Mínimo de Horas Relógio Cursadas" in requisito:
            n = requisito.split("= ")[-1]
            if n != "0":
                resultados.append("RP" + str(n) + "HR")

    separador = ","
    return separador.join(resultados)


Tk().withdraw()
arquivo_excel = askopenfilename(
    filetypes=[('Arquivo excel', '.xlsx')],
    title='Selecione o arquivo excel')
arquivo_word = askopenfilename(
    filetypes=[('Arquivo excel', '.xlsx')],
    title='Selecione o arquivo word')

# arquivo_excel = "Engenharia Producao Grade_curricular_CEPD 2018_1-M.xlsx"
# arquivo_word = "Re 011-2020 Consun Ad. ref. Alt. MC Eng. de Produção 2018.xlsx"

arquivo_de_saida = "Relatório de Erros.xlsx"
saida_workbook = load_workbook(arquivo_de_saida)
saida_workbook.remove(saida_workbook.active)
saida_workbook.create_sheet(title="Relatório")
saida_sheet = saida_workbook.worksheets[0]

workbook_excel = xlrd.open_workbook(arquivo_excel)
sheet_excel = workbook_excel.sheet_by_name("Ciclos")

workbook_word = xlrd.open_workbook(arquivo_word)
sheet_word = workbook_word.sheet_by_name("Plan1")

disciplinas_excel = []
disciplinas_word = []

# Dados Word
ciclo = 0
ordem_word = []

for linha in range(0, sheet_word.nrows):

    if type(sheet_word.cell_value(linha, 0)) is str:
        if "PERÍODO" in str.upper(sheet_word.cell_value(linha, 0)):
            ciclo = int(
                sheet_word.cell_value(linha, 0).replace("°", "|").replace("º", "|").replace(".", "").split("|")[0])
        continue
    ordem_disciplina = int(sheet_word.cell_value(linha, 0))
    nome_disciplina = sheet_word.cell_value(linha, 1).strip().replace("–", "-")
    requisito_disciplina = sheet_word.cell_value(linha, 2).replace(" ", "").replace("\xa0", "").replace(";", ",")\
        .strip("0")
    aula_teorica = int(sheet_word.cell_value(linha, 3))
    aula_pratica = int(sheet_word.cell_value(linha, 4))
    creditos = int(sheet_word.cell_value(linha, 5))
    ch_oficial = int(sheet_word.cell_value(linha, 6))
    ch_relogio_oficial = int(sheet_word.cell_value(linha, 7))

    disciplinas_word.append([
        ciclo, nome_disciplina, requisito_disciplina, aula_teorica,
        aula_pratica, creditos, ch_oficial, ch_relogio_oficial])
    ordem_word.append(ordem_disciplina)

# Dados Excel
ciclo = 0

for linha in range(0, sheet_excel.nrows):
    if sheet_excel.cell_value(linha, 1) == "" or \
            sheet_excel.cell_value(linha, 1) == "Total Carga Horária (mínimo):" or \
            sheet_excel.cell_value(linha, 1) == "Disciplina" or \
            sheet_excel.cell_value(linha, 1) == "Total":
        continue
    if sheet_excel.cell_value(linha, 1) == "Ciclo:":
        ciclo = int(sheet_excel.cell_value(linha, 2))
        continue

    nome_disciplina = sheet_excel.cell_value(linha, 1).strip().split(" - ")[1:]
    separador_nome = " – "
    nome_disciplina = separador_nome.join(nome_disciplina).strip().replace("–", "-")

    requisito_disciplina = tratar_requisitos(sheet_excel.cell_value(linha, 17))
    aula_teorica = int(sheet_excel.cell_value(linha, 6) / 20)
    aula_pratica = int(sheet_excel.cell_value(linha, 7) / 20)
    creditos = int(sheet_excel.cell_value(linha, 15))
    ch_oficial = int(sheet_excel.cell_value(linha, 8))
    ch_relogio_oficial = int(sheet_excel.cell_value(linha, 12))

    disciplinas_excel.append(
        [ciclo, nome_disciplina, requisito_disciplina, aula_teorica, aula_pratica, creditos, ch_oficial,
         ch_relogio_oficial])

print("\n----Excel----")
for i in disciplinas_excel:
    print(i)

print("\n----Word----")
for i in disciplinas_word:
    print(i)

consistente = 0
inconsistente = 0

# Header
saida_sheet.append(["Coluna(s) com erro", "Ordem", "Ciclo", "Disciplina", "Requisitos", "Carga horária teórica",
                    "Carga horária prática", "Créditos", "Carga Horária oficial", "Carga Horária relógio oficial"])

print("\n----Relatório----")
for index, word in enumerate(disciplinas_word):
    if "Eletiva" in word:
        continue
    elif word in disciplinas_excel:
        consistente += 1
    else:
        erro = []
        existe_no_excel = any(word[1] in sublist for sublist in disciplinas_excel)
        if existe_no_excel:
            indice = [x for x in disciplinas_excel if word[1] in x][0]
            excel = disciplinas_excel[disciplinas_excel.index(indice)]
            if word[0] != excel[0]:
                erro.append("ciclo")
            if word[2] != excel[2]:
                erro.append("requisitos")
            if word[3] != excel[3]:
                erro.append("carga horária teórica")
            if word[4] != excel[4]:
                erro.append("carga horária prática")
            if word[5] != excel[5]:
                erro.append("créditos")
            if word[6] != excel[6]:
                erro.append("carga horária oficial")
            if word[7] != excel[7]:
                erro.append("carga horária relógio oficial")
        else:
            erro.append("disciplina")
        separador_nome = ","
        erros = separador_nome.join(erro)
        word.insert(0, ordem_word[index])
        word.insert(0, erros)
        print(word)
        saida_sheet.append(word)
        inconsistente += 1

print("\nValores consistentes: " + str(consistente))
print("Valores inconsistentes: " + str(inconsistente))
print("Alerta: " + str(alerta))

saida_workbook.save(arquivo_de_saida)
saida_workbook.close()

print("\n" + arquivo_word
      .replace("C:/Users/Vinicius/Grupo Marista/DAC - Compliance e Informações Acadêmicas - Ação Matrizes 2020-1/", "")
      .replace("Novo(a) Planilha do Microsoft Excel.xlsx", ""))
